import openpyxl
import xlrd
import GlobalVariables
import ReadXML
import os


def create_files_from_excel(xl_location, file_to_fill_names,  save_to_location, save_to_server_location):
    try:
        workbook = xlrd.open_workbook(xl_location)
        f_list_names = open(file_to_fill_names, 'w+')

        for sheet in workbook.sheets():
            for row in range(sheet.nrows):
                if row == 0:
                    continue
                name = sheet.cell(row, 1).value
                name = name[:name.find('.testplan')]

                protocol = sheet.cell(row, 2).value
                protocol = protocol[protocol.find('Automation_Test_plans/'):len(protocol)]
                protocol = protocol[protocol.find('/'):].replace('/', '')
                giga = sheet.cell(row, 6).value
                status = sheet.cell(row, 8).value
                if 'finish' in status or 'not to run' in status:
                    continue

                f_list_names.write(protocol + ',' + name + ',' + str(giga) + ',' + '\n')

                cmd = sheet.cell(row, 4).value

                if not os.path.exists(save_to_location + protocol + '/'):
                    os.makedirs(save_to_location + protocol + '/')

                path_to_file = save_to_location + protocol + '/' + name + '.txt'

                f = open(path_to_file, 'w+')
                f.write(cmd)
                f.close()

                # if test need to run with Server side in DUT
                if ('Client' in name or 'client' in name) and 'IEEE1588' not in name:
                    server_cmd = sheet.cell(row, 5).value
                    server_test_name = name.replace('Client', 'Server')
                    server_protocol = protocol.replace('Client', 'Server')

                    if not os.path.exists(save_to_server_location + server_protocol + '/'):
                        os.makedirs(save_to_server_location + server_protocol + '/')

                    path_to_server_file = save_to_server_location + server_protocol + '/' +\
                                          server_test_name + '_For_Client.txt'

                    f = open(path_to_server_file, 'w+')
                    f.write(server_cmd)
                    f.close()

        f_list_names.close()
    except Exception as e:
        GlobalVariables.logger.error(e.message)


def read_status_test(xl_location, test_name):
    result = ''
    workbook = openpyxl.load_workbook(xl_location)

    for sheet in workbook.worksheets:
        for row in range(sheet.max_row + 1):
            if row == 0 or row == 1:
                continue
            name = sheet.cell(row, 2).value
            if name is not None and 'testplan' in name:
                name = name[:name.find('.testplan')]
                if name == test_name:
                    result = sheet.cell(row, 9).value

    return result


def check_result(dic, str_status):
    err_msg = 'ERROR OCCURRED in update XL- conflict.'

    if "stuck!!" in dic['state']:
        if 'PASS' in str_status or 'FAIL' in str_status:
            err_msg += "\nThe state is STUCK, not PASS or FAIL."
            return err_msg
        return str_status

    # taking numbers of remaining cases to check if test ===really=== end
    num_remaining = dic['remaining-cases']
    # taking numbers of failed cases to check if test is ===really=== fail
    num_fail = dic['failed-cases']
    err_msg += " \nremaining cases are not 0."

    if 'FAIL' in str_status:
        if num_remaining == 0:
            if num_fail == 0:
                return 'finished!! - PASS'
            else:
                return str_status
        else:
            return err_msg

    if 'PASS' in str_status:
        if num_remaining == 0:
            return str_status
        else:
            return err_msg

    return str_status


def update_test_status_in_excel(xl_location, testplan_name, status, time, log_dir):
    try:
        GlobalVariables.excel_lock.acquire()
        found = 0
        workbook = openpyxl.load_workbook(xl_location)
        for sheet in workbook.worksheets:
            for row in range(sheet.max_row + 1):
                if row == 0 or row == 1:
                    continue
                name = sheet.cell(row, 2).value
                if name is not None and 'testplan' in name:
                    name = name[:name.find('.testplan')]
                    if name == testplan_name:
                        found = 1
                        # check if the result i got is appropriate to run-time-info.xml results
                        dic = ReadXML.read_xml_results(log_dir + '/run-time-info.xml', name, '', '', time)
                        status = check_result(dic, status)

                        GlobalVariables.logger.info("found the finished test in excel table")
                        sheet.cell(row, 9).value = status
                        if "reproduce_failures" not in log_dir:
                            sheet.cell(row, 10).value = (str(time))
                        sheet.cell(row, 11).value = GlobalVariables.LP_machine
                        GlobalVariables.logger.info(name + " " + sheet.cell(row, 9).value + " - updated XL")
                        break

        if found == 0:
            GlobalVariables.logger.info("didn't find the finished test in excel table!")

        workbook.save(xl_location)
        GlobalVariables.excel_lock.release()
    except Exception as e:
        GlobalVariables.logger.error(e.message)
