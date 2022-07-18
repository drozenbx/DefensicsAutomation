#!/usr/bin/env python
import os
activate_this = os.path.expanduser("~/PycharmProjects/DefensicsAutomation/venv/bin/activate_this.py")
execfile(activate_this, dict(__file__=activate_this))

import xlrd
import xlsxwriter
import openpyxl


# import GlobalVariables
# import XL_actions

#path_cycle_result = 'root/synopsys/defensics/Automation_Cycle_Results/' + GlobalVariables.project + '.xlsx'
#
path_testplans_xl = '/root/PycharmProjects/DefensicsAutomation/FilesForAutomation/files/testplans_xl.xlsx'
path_cycle_xl = '/root/PycharmProjects/DefensicsAutomation/FilesForAutomation/files/cycle_testplans.xlsx'


def prepares_cycle_xl():


#cannt install SHUTIL
    #shutil.copy(path_full_test_plans_xl, path_cycle_test_plans_xl)

    testplans_xl = openpyxl.load_workbook(path_testplans_xl)
    sheets_testplans_xl = testplans_xl.sheetnames
    sheet_testplans_xl = testplans_xl[sheets_testplans_xl[0]]

    cycle_xl = openpyxl.load_workbook(path_cycle_xl)
    sheets_cycle_xl = cycle_xl.sheetnames
    sheet_cycle_xl = cycle_xl[sheets_cycle_xl[0]]

# loop - testplans_xl
    # for sheet_xl in testplans_xl.sheets():
    row_xl = 2
    while sheet_testplans_xl.cell(row_xl, 1).value:
        test_name = sheet_testplans_xl.cell(row_xl, 1).value

        exist_test = 0
        # loop - list_cycle_testplans_xl
        row_cycle = 1
        while sheet_cycle_xl.cell(row_cycle, 1).value:
            print 'testplan xl test' + test_name
            print 'rc', row_cycle
            print 'rt', row_xl
            print 'cycle xl test', sheet_cycle_xl.cell(row_cycle, 1).value
            if sheet_cycle_xl.cell(row_cycle, 1).value == test_name:
                exist_test = 1
                break
            row_cycle += 1

        if exist_test == 1:
            print 'exist_test == 1'
            while sheet_testplans_xl.cell(row_xl, 1).value == test_name:
                sheet_testplans_xl.cell(row_xl, 8).value = ''
                sheet_testplans_xl.cell(row_xl, 9).value = ''
                sheet_testplans_xl.cell(row_xl, 10).value = ''
                row_xl += 1
            # exist_test = 1
            break

        else:
            print 'exist_test == 0'
            while sheet_testplans_xl.cell(row_xl, 1).value == test_name:
                sheet_testplans_xl.cell(row_xl, 8).value = 'not to run'
                sheet_testplans_xl.cell(row_xl, 9).value = ''
                sheet_testplans_xl.cell(row_xl, 10).value = ''
                row_xl += 1

    testplans_xl.save(path_testplans_xl)
    print testplans_xl.save(path_testplans_xl)
    cycle_xl.save(path_cycle_xl)

# workbook = xlsxwriter.Workbook(path_to_xl_cycle)
# header_line = []
#
# for sheet in workbook.sheets():
#     pre_test_name = sheet.cell(0, 0).value
#     # nrows- all of rows in sheet
#     for row in range(sheet.nrows):
#
#         if row == 0:
#             for col in range(sheet.ncols):
#                 # get header line to add to start of file in a list
#                 header_line.append(sheet.cell(0, col).value)
#             continue
#
#         test_name = sheet.cell(row, 0).value
#         if pre_test_name in test_name
#
#             XL_actions.write_row(path_to_xl_cycle, ,)
#             pre_test_name = test_name


if __name__ == '__main__':
    print('start')
    prepares_cycle_xl()
    print('finished')
