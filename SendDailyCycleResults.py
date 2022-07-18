#!/usr/bin/env python
# coding=utf-8
import os
activate_this = os.path.expanduser("~/PycharmProjects/DefensicsAutomation/venv/bin/activate_this.py")
execfile(activate_this, dict(__file__=activate_this))


from datetime import datetime
import GlobalVariables
import GlobalFunctions
import email_utils
import subprocess
import openpyxl
import xlrd


global_path_to_files = '/root/PycharmProjects/DefensicsAutomation/FilesForAutomation/manage_cycle_results/'
path_to_list_setups = global_path_to_files + 'running_setups.txt'
path_to_excel_to_update_cycle = global_path_to_files + 'excel_to_update_cycle.xlsx'
path_to_cycle_start_time = global_path_to_files + 'cycle_start_time.txt'
path_to_dir_remote_setups_results = global_path_to_files + 'remote_setups_results/'
path_to_dir_of_protocols_Xls = global_path_to_files + 'protocols_part_to_XLs/'
path_to_list_all_full_tests_for_cycle = global_path_to_files + 'list_all_full_tests_for_cycle.txt'
print(path_to_list_all_full_tests_for_cycle)
all_status = {'pass': 0, 'fail': 0, 'error': 0, 'running': 0, 'no run': 0, 'num_tests': 0}
list_running_setups_name = []
list_running_setups_ip = []
XL_name = ''
str_pass = 'pass'
str_fail = 'fail'
str_error = 'error'
str_running = 'running'
str_no_run = 'no run'
all_run_time = '--:--:--'
send_daily = 0
check_big_XL = False


def copy_XL_from_remote_running_setups():
    """
    This function copy XL results from setup list locally
    """
    global list_running_setups_name
    global list_running_setups_ip
    global XL_name

    # reading running setups from file
    # פותח את RUNNING_SETUPS
    fp = open(path_to_list_setups, 'rb')

    for setup in fp.read().splitlines():
        setup_ip = setup[setup.find('-') + 1:]
        setup_name = setup[:setup.find('-')]
        #  רשימת השמות של המכונות שמריצות את הסייקל
        list_running_setups_name.append(setup_name)
        #  רשימת הIP של המכונות שמריצות את הסייקל
        list_running_setups_ip.append(setup_ip)
        print(setup_ip)

        # copy remote XL of setup using 'net' command
        subprocess.call('cp -rf /net/{0}{1} {2}'.format(setup_ip, GlobalVariables.path_to_XL,
                                                        path_to_dir_remote_setups_results), shell=True, stderr=subprocess.PIPE)
        XL_name = GlobalVariables.path_to_XL[GlobalVariables.path_to_XL.rfind('/') + 1:]
        print(XL_name)
        print("!!!!!!!!!!!!!!!!!!!")
        print(GlobalVariables.path_to_XL)
        # if XL was copied successfully, add to name the setup name
        print(path_to_dir_remote_setups_results + XL_name)
        if os.path.isfile(path_to_dir_remote_setups_results + XL_name):
            subprocess.call('mv {0} {1}'.format(path_to_dir_remote_setups_results + XL_name,
                                                path_to_dir_remote_setups_results + setup_ip + '_' + XL_name), shell=True, stderr=subprocess.PIPE)
        else:
            GlobalVariables.logger.error("XL results file was not copied correctly from setup " + setup_ip +
                                         '\nPlease check connection and path to the file.\nSearching file: ' + path_to_dir_remote_setups_results + XL_name)
            exit(0)


def update_status_of_part_run(path, name_part_test, status, time_run, setup_name):
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets[0]

    for row in range(sheet.max_row):
        if row == 0:
            continue

        current_name = sheet.cell(row + 1, 2).value
        current_name = current_name[:current_name.find('.testplan')]

        # save current status in the table- if it was already pass wont change it
        if sheet.cell(row + 1, 9).value is None:
            current_status = ''
        else:
            current_status = sheet.cell(row + 1, 9).value

        # check if part test is like the current test- then if yes update status and time
        if current_name == name_part_test:
            if 'FAIL' in current_status:
                if 'PASS' in status:
                    sheet.cell(row=row + 1, column=9).value = status
                    sheet.cell(row=row + 1, column=10).value = time_run
                    sheet.cell(row=row + 1, column=11).value = setup_name

            elif 'PASS' not in current_status and status != '' and status is not None:
                sheet.cell(row=row + 1, column=9).value = status
                sheet.cell(row=row + 1, column=10).value = time_run
                sheet.cell(row=row + 1, column=11).value = setup_name
            break

    wb.save(path)


def open_and_update_XLs(path_xl, setup_name):
    try:
        workbook = xlrd.open_workbook(path_xl)

        for sheet in workbook.sheets():
          for row in range(sheet.nrows):
            if row == 0:
                continue

            protocol_and_test_name = sheet.cell(row, 0).value

            status = sheet.cell(row, 8).value
            if 'not to run' in status:
                continue

            name = sheet.cell(row, 1).value
            if name is None or name == '':
                continue

            name = name[:name.find('.testplan')]
            time_run = sheet.cell(row, 9).value

            path_to_file = path_to_dir_of_protocols_Xls + protocol_and_test_name + '.xlsx'

            if not os.path.isfile(path_to_file):
                continue

            if send_daily == 1:
                GlobalVariables.logger.info(name + ' ' + status)
            elif send_daily == 0:
                setup_name = sheet.cell(row, 10).value

            # update protocol+name XL
            update_status_of_part_run(path_to_file, name, status, time_run, setup_name)
    except Exception as E:
        print(path_xl)


def update_protocols_XLs_locally():
    global list_running_setups_name
    global list_running_setups_ip
    global XL_name
    global send_daily
    global check_big_XL
    i = 0

    if send_daily == 1:
        GlobalVariables.logger.info("Collect results from the Setups: ")

    # run over all the running setups in the list
    for setup in list_running_setups_ip:
        setup_name = list_running_setups_name[i]
        i += 1

        if send_daily == 1:
            GlobalVariables.logger.info('\n' + str(i) + ') ' + setup_name)

        path_xl = path_to_dir_remote_setups_results + setup + '_' + XL_name

        open_and_update_XLs(path_xl, setup_name)

    # run over the BIG XL (if exist) to update cycle results
    # name of excel: 'excel_to_update_cycle.xlsx'
    if check_big_XL is True:
        GlobalVariables.logger.info("\nChecking big XL.........")
        send_daily = 0
        open_and_update_XLs(path_to_excel_to_update_cycle, '')#GlobalVariables.server_machine)


def add_run_time(old_run_time, additional_run_time, name):
    if old_run_time == "--:--:--":
        if additional_run_time != '' and additional_run_time is not None and additional_run_time \
                != "--:--:--" and additional_run_time != "00:00:00":
            return additional_run_time
        return old_run_time

    if additional_run_time == '' or additional_run_time is None or additional_run_time \
            == "--:--:--" or additional_run_time == "00:00:00":
        return old_run_time

    new_time = GlobalFunctions.sum_time(old_run_time, additional_run_time, name)
    return new_time


def get_summarized_line_to_test(name, num_parts):
    global all_status
    global all_run_time

    num_parts = int(num_parts)
    path_to_file = path_to_dir_of_protocols_Xls + name + '.xlsx'
    test_status = {'pass': 0, 'fail': 0, 'error': 0, 'running': 0, 'no run': 0}
    run_time = '--:--:--'
    html_lime = ''
    setups_ran_protocol = ""
    all_status['num_tests'] += 1

    wb = openpyxl.load_workbook(path_to_file)
    sheet = wb.worksheets[0]

    for row in range(sheet.max_row):
        if row == 0:
            continue

        status = sheet.cell(row=row + 1, column=9).value
        part_run_time = sheet.cell(row=row + 1, column=10).value
        run_time = add_run_time(run_time, part_run_time, name)

        if sheet.cell(row=row + 1, column=11).value is not None and sheet.cell(row=row + 1, column=11).value is not ''\
                and sheet.cell(row=row + 1, column=11).value not in setups_ran_protocol:
            setups_ran_protocol += sheet.cell(row=row + 1, column=11).value + ', '

        # check status part of test
        if status == '' or status is None:
            test_status[str_no_run] += 1
        # TODO also skip are pass???
        elif 'PASS' in status or 'SKIP' in status:
            test_status[str_pass] += 1
        elif 'FAIL' in status:
            test_status[str_fail] += 1
        elif 'RUNNING' in status:
            test_status[str_running] += 1
        elif 'ERROR' in status or 'WARNING' in status:
            test_status[str_error] += 1
        else:
            # GlobalVariables.logger.info('status->>> ' + status)
            test_status[str_no_run] += 1

    # add html line of one test
    html_lime += '\n<tr><td>{}.</td><td>{}</td><td align="right">{}</td>' \
        .format(all_status['num_tests'], name, run_time)

    # check test global status and add to line
    if test_status[str_pass] == num_parts:
        all_status[str_pass] += 1
        html_lime += '<td bgcolor="green"><font color="white">PASS</font></td>'
    elif test_status[str_no_run] == num_parts:
        all_status[str_no_run] += 1
        html_lime += '<td bgcolor="gray"><font color="white">NO RUN</font></td>'
    elif test_status[str_no_run] == num_parts:
        all_status[str_no_run] += 1
        html_lime += '<td bgcolor="gray"><font color="white">NO RUN</font></td>'
    elif test_status[str_running] > 0 or test_status[str_no_run] > 0:
        all_status[str_running] += 1
        html_lime += '<td bgcolor="blue"><font color="white">RUNNING</font></td>'
    elif test_status[str_running] == 0 and test_status[str_no_run] == 0 and test_status[str_error] > 0:
        all_status[str_error] += 1
        html_lime += '<td bgcolor="black"><font color="white">ERROR</font></td>'
    elif test_status[str_running] == 0 and test_status[str_no_run] == 0 and test_status[str_fail] > 0:
        all_status[str_fail] += 1
        html_lime += '<td bgcolor="red"><font color="white">FAIL</font></td>'
    # TODO what is the status if nothing is right?
    else:
        all_status[str_no_run] += 1
        html_lime += '<td bgcolor="gray"><font color="white">NO RUN</font></td>'

    # add percent status in html line
    pass_percent = str(int((test_status[str_pass] * 100.0) / num_parts)) + '%'
    fail_percent = str(int((test_status[str_fail] * 100.0) / num_parts)) + '%'
    error_percent = str(int((test_status[str_error] * 100.0) / num_parts)) + '%'
    running_percent = str(int((test_status[str_running] * 100.0) / num_parts)) + '%'
    no_run_percent = str(int((test_status[str_no_run] * 100.0) / num_parts)) + '%'

    # setups_ran_protocol = setups_ran_protocol[len(setups_ran_protocol):]

    html_lime += '<td align="right">{0}</td><td align="right">{1}</td><td align="right">{2}</td>' \
                 '<td align="right">{3}</td><td align="right">{4}</td><td align="right">{5}</td></tr>'\
        .format(pass_percent, fail_percent, error_percent, running_percent, no_run_percent, setups_ran_protocol)

    if run_time is not "--:--:--":
        old = all_run_time
        # add the test run time to the global varuable a all tests run time
        all_run_time = add_run_time(all_run_time, run_time, name)

    return html_lime


def collect_all_results_by_protocol():
    table_lines = ''

    # reading list of all full tests from file
    with open(path_to_list_all_full_tests_for_cycle, 'rb') as fp:
        list_full_tests = fp.read().splitlines()

    # sort the list by name test
    list_full_tests.sort()

    for test in list_full_tests:
        name = test[:test.find(':')]
        num_parts_for_test = test[test.find(': ') + 2:]

        table_lines += get_summarized_line_to_test(name, num_parts_for_test)
    return table_lines


def create_html(lines):
    global all_run_time
    date = str(datetime.now())
    date = date.replace(' ', '_')
    date = date.replace(':', '_')

    # Calculates the actual cycle time from the start of the cycle
    with open(path_to_cycle_start_time, 'r') as file:
        start_cycle_time = file.readline()
    start_cycle_time = start_cycle_time[start_cycle_time.find('=') + 1:start_cycle_time.find('|')]
    start_date_time_obj = datetime.strptime(start_cycle_time, '%Y-%m-%d %H:%M:%S.%f')
    current_cycle_time = GlobalFunctions.get_time(start_date_time_obj)
    name_result = date + '_' + GlobalVariables.project + '_cycle'
    path_to_html = GlobalVariables.path_to_cycle_results + name_result + '.html'

    if not os.path.exists(GlobalVariables.path_to_cycle_results):
        os.makedirs(GlobalVariables.path_to_cycle_results)
    print("hhh")
    with open(path_to_html, 'w') as html_file:
        print(path_to_html)
        html_file.write('<html><head>')
        print("???")
        # link the javascript file to your html using the <script> tag
        html_file.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/moment.js/2.10.1/moment.min.js">'
                        '</script><script src="https://ajax.googleapis.com/ajax/libs/jquery/3.3.1/jquery.min.js">'
                        '</script>')
        print("!!!")
        html_file.write('<title>Defensics SAFE Cycle Status</title>')
        html_file.write('</head><body><h1>')
        html_file.write('Defensics SAFE Cycle Status')
        html_file.write('</h1>\n\n<h3>(Title template: PASS | FAIL | ERROR | RUNNING | NO RUN)')
        html_file.write('</h3>\n\n<pre>')
        html_file.write("\nRunning Time of cycle is: " + str(current_cycle_time))
        # print(all_run_time)
        # TODO   html_file.write("\nSum time of all runs is: " + all_run_time)

        html_file.write("\nMachines name (LP): ")

        setups = ''
        for setup in list_running_setups_name:
            setups += setup + ', '
            print(list_running_setups_name)
        if setups is not '':
            html_file.write(setups[:setups.rfind(',')])

        html_file.write("\nResults path: " + GlobalVariables.LP_machine + ': ' + path_to_html)
        #table
        html_file.write('</pre>\n\n<br /><br /><table><tr>\n')
        html_file.write('<td bgcolor="gray"><font color="white"># &nbsp&nbsp&nbsp&nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">Protocol &nbsp/&nbsp Name Test '
                        '&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp</font></td>')
        html_file.write('<td bgcolor="gray" align="right"><font color="white">Time &nbsp&nbsp&nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">Status &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">% Pass &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">% Fail &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">% Error &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">% Running &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">% No Run &nbsp</font></td>')
        html_file.write('<td bgcolor="gray"><font color="white">Setups &nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp'
                        '&nbsp&nbsp&nbsp&nbsp</font></td>')
        html_file.write('</tr>\n')

        html_file.write(lines)
        html_file.write('</table><br /><br />')
        html_file.write('')

        pass_percent = str(int((all_status[str_pass] * 100.0) / all_status['num_tests']))
        #print("all pass" + str(int((all_status[str_pass]))))
        print("percent pass" + pass_percent)
        print('all status' + str(all_status))
        fail_percent = str(int((all_status[str_fail] * 100.0) / all_status['num_tests']))
        error_percent = str(int((all_status[str_error] * 100.0) / all_status['num_tests']))
        running_percent = str(int((all_status[str_running] * 100.0) / all_status['num_tests']))
        no_run_percent = str(int((all_status[str_no_run] * 100.0) / all_status['num_tests']))
        print("html_file")
        html_file.write("\n<table><tr><td>number of tests passed " + str(all_status[str_pass]) + "</td><td>[ "
                        + str(pass_percent) + "% PASS ]</td></tr>\n")
        html_file.write(
            "\n<tr><td>number of tests fail " + str(all_status[str_fail]) + "</td><td>[ " + str(fail_percent)
            + "% FAIL ]</td></tr>\n")
        html_file.write("\n<tr><td>number of tests error " + str(all_status[str_error]) + "</td><td>[ " + str(
            error_percent) + "% ERROR ]</td></tr>\n")
        html_file.write("\n<tr><td>number of tests running " + str(all_status[str_running]) + "</td><td>[ " + str(
            running_percent) + "% RUNNING ]</td></tr>\n")
        html_file.write("\n<tr><td>number of tests no run " + str(all_status[str_no_run]) + "</td><td>[ " + str(
            no_run_percent) + "% NO RUN ]</td></tr>\n")
        html_file.write("\n<tr><td colspan=\"2\">----------------------------------</td></tr>\n")
        html_file.write("\n<tr><td>number of tests: </td><td>" + str(all_status['num_tests']) + "</td></tr>\n")
        html_file.write("\n\t<tr><td>number of finished tests: </td><td>" + str(all_status[str_pass] + all_status[str_fail] + all_status[str_error]) + "</td></tr>")
        html_file.write("\n\t<tr><td>number of running tests: </td><td>" + str(all_status[str_running]) + "</td></tr>")
        html_file.write("\n\t<tr><td>number of waiting tests: </td><td>" + str(all_status[str_no_run]) +
                        "</td></tr></table>")
        html_file.write('</body></html>')
    print("end create_html")
    return path_to_html


def send_email_results(path_to_html, additional_txt_tittle):
    # try:
    # path_to_html = GlobalVariables.path_to_cycle_results + "/2019-01-21_18_11_06.374425_cvl_cycle_Shlomi3.html"
    GlobalVariables.logger.info("\nPublish Results....")
    
    recipients_file = GlobalVariables.path_to_recipients_file
    #print(path_to_recipients_file)
    # taking numbers of success/ failures/ running/ error/ no run tests
    num_passed = all_status[str_pass]
    num_failed = all_status[str_fail]
    num_running = all_status[str_running]
    num_error = all_status[str_error]
    num_no_run = all_status[str_no_run]

    # reading recipients list from file
    with open(recipients_file, 'rb') as fp:
        recipients_list = fp.read().splitlines()

    email_utils.email_to = recipients_list

    quick_results = GlobalVariables.project + ' - ' + str(num_passed) + ' | ' \
                    + str(num_failed) + ' | ' + str(num_error) + ' | ' + str(num_running) + ' | ' + str(num_no_run)\
                    + ' -> ' + additional_txt_tittle

    message = email_utils.create_message(quick_results)

    with open(path_to_html) as f:
        html = f.read()
    email_utils.set_body(message, html)
    print("------------------")
    email_utils.send_email(message, recipients_list)
    print("------------------")
    GlobalVariables.logger.info("Successfully sent:)")

    # except Exception as e:
    #     GlobalVariables.logger.error(e.message)
    #     print(e)
    #     print(e.message)


if __name__ == '__main__':
    send_daily = 1
    additional_txt_tittle = "Daily Cycle Results"
    #מנתב לרימוט
    GlobalVariables.path_to_lists_runs_details = global_path_to_files
    GlobalFunctions.enable_logging('')
    copy_XL_from_remote_running_setups()
    update_protocols_XLs_locally()
    table_lines = collect_all_results_by_protocol()
    path_to_html = create_html(table_lines)
    send_email_results(path_to_html, additional_txt_tittle)
