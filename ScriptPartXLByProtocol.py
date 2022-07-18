#!/usr/bin/env python
import os
activate_this = os.path.expanduser("~/PycharmProjects/DefensicsAutomation/venv/bin/activate_this.py")
execfile(activate_this, dict(__file__=activate_this))

import openpyxl
import xlrd
import xlsxwriter
import os
import subprocess
from datetime import datetime


global_path_to_files = '/root/PycharmProjects/DefensicsAutomation/FilesForAutomation/manage_cycle_results/'
path_to_full_cycle_XL = global_path_to_files + \
                        "AutomationChecks_LKV.xlsx"
path_to_dir_of_protocols_Xls = global_path_to_files + 'protocols_part_to_XLs/'
path_to_list_setups = global_path_to_files + 'running_setups.txt'
path_to_cycle_start_time = global_path_to_files + 'cycle_start_time.txt'
path_to_dir_remote_setups_results = global_path_to_files + 'remote_setups_results/'
dic_list_protocol_with_name_test = {}


def write_row(path, new_line, row):
    """
    This function write a new row into the XL file (path)

    :param path: full path to the XL file
    :param new_line: new line to add in- formatted like a list
    :param row: num of the row we want to add
    """
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    for column, heading in enumerate(new_line):
        sheet.cell(row=row + 1, column=column + 1).value = heading  # header_line[0]
    wb.save(path)


def split_part_runs_by_protocol_to_xls():
    """
    This function split a big XL of parted protocols to XLs that will collect all parts of protocol in 1 XL file.
    """
    global list_protocol_with_name_test

    workbook = xlrd.open_workbook(path_to_full_cycle_XL)
    # f_list_names = open(file_to_fill_names, 'w+')
    header_line = []

    for sheet in workbook.sheets():
        for row in range(sheet.nrows):
            new_line_test = []

            if row == 0:
                for col in range(sheet.ncols):
                    # get header line to add to start of file in a list
                    header_line.append(sheet.cell(0, col).value)
                continue

            status = sheet.cell(row, 8).value
            if 'not to run' in status:
                continue

            protocol_and_test_name = sheet.cell(row, 0).value

            path = path_to_dir_of_protocols_Xls + protocol_and_test_name + '.xlsx'

            # if XL doesn't exist create one to protocol
            if not os.path.isfile(path):
                workbook = xlsxwriter.Workbook(path)
                worksheet = workbook.add_worksheet()
                workbook.close()

                # add the protocol with name test to the dic list tests
                dic_list_protocol_with_name_test[protocol_and_test_name] = 1
            else:
                dic_list_protocol_with_name_test[protocol_and_test_name] += 1

                # get new line to add to file in a list
            for col in range(sheet.ncols):
                new_line_test.append(sheet.cell(row, col).value)

            workbook_1 = xlrd.open_workbook(path)
            # TODO work on few sheet??
            sheet_1 = workbook_1.sheet_by_name('Sheet1')

            # if the file is empty
            if sheet_1.nrows == 0:
                write_row(path, header_line, 0)
                write_row(path, new_line_test, 1)
            # else- the header already exist
            else:
                write_row(path, new_line_test, sheet_1.nrows)


def check_dir_is_empty():
    if os.listdir(path_to_dir_of_protocols_Xls):
        print("The folder '" + path_to_dir_of_protocols_Xls + "' is not empty. \nPlease clear it before running the script")
        exit(0)


def create_file_with_protocol_and_name():
    global list_protocol_with_name_test

    with open(global_path_to_files + '/list_all_full_tests_for_cycle.txt', 'w+') as file_list:
        for test in dic_list_protocol_with_name_test.keys():
            file_list.write(test + ': ' + str(dic_list_protocol_with_name_test[test]) + '\n')


def update_start_time_of_cycle():
    """
    This function update the start time of the new cycle
    """

    subprocess.call(['sed', '-i', '/' + 'start' + ',/d', path_to_cycle_start_time])

    my_file = open(path_to_cycle_start_time, 'w+')
    start_time = datetime.now()
    my_file.write('start=' + str(start_time) + '|')
    my_file.close()
    print("Start time of Cycle was updated.")


if __name__ == '__main__':
    check_dir_is_empty()
    print('Working..... \nplease wait :)')
    split_part_runs_by_protocol_to_xls()
    create_file_with_protocol_and_name()
    update_start_time_of_cycle()
    print('Finished to part big XL to protocols XL.')
    print('A file contains list of the tests to run in this Cycle was created.')
